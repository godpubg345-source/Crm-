"""
Import courses from Excel file into database.

Usage:
    python manage.py import_courses --file=courses.xlsx
    python manage.py import_courses --file=courses.xlsx --university="University of Manchester"
    python manage.py import_courses --generate-template
"""
from django.core.management.base import BaseCommand
from datetime import datetime
from decimal import Decimal
import pandas as pd
from universities.models import University, Course, CourseRequirement


class Command(BaseCommand):
    help = 'Import courses from Excel file or generate template'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Path to Excel file containing course data'
        )
        parser.add_argument(
            '--university',
            type=str,
            help='Filter to specific university name'
        )
        parser.add_argument(
            '--generate-template',
            action='store_true',
            help='Generate empty Excel template'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview import without saving'
        )

    def handle(self, *args, **options):
        if options['generate_template']:
            self.generate_template()
            return

        if not options['file']:
            self.stderr.write('Please provide --file or --generate-template')
            return

        self.import_courses(
            options['file'],
            options.get('university'),
            options.get('dry_run', False)
        )

    def generate_template(self):
        """Generate Excel template for course data entry."""
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Main courses sheet
        ws = wb.active
        ws.title = "Courses"
        
        headers = [
            'university_name', 'course_name', 'level', 'duration', 'duration_months',
            'tuition_fee_intl', 'tuition_fee_home', 'currency', 'deposit_required',
            'intake_january', 'intake_may', 'intake_september', 'intakes',
            'ielts_overall', 'ielts_each_band', 'academic_requirement',
            'work_experience_required', 'work_experience_years',
            'official_url', 'department', 'course_code'
        ]
        
        # Header style
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Add sample data row
        sample_row = [
            'University of Manchester', 'MSc Data Science', 'PG', '1 Year', 12,
            29000, 15000, 'GBP', 2000,
            'Yes', 'No', 'Yes', 'Sep, Jan',
            6.5, 6.0, '2:1 degree in related field',
            'No', 0,
            'https://www.manchester.ac.uk/study/masters/data-science', 'Computer Science', 'CS-101'
        ]
        for col, value in enumerate(sample_row, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Country requirements sheet
        ws2 = wb.create_sheet("Country_Requirements")
        country_headers = [
            'course_name', 'university_name', 'country_code',
            'min_qualification', 'accepted_qualifications', 'min_gpa', 'min_percentage', 'min_cgpa',
            'ielts_overall', 'ielts_speaking', 'ielts_writing', 'ielts_reading', 'ielts_listening',
            'toefl_score', 'pte_score', 'duolingo_score',
            'work_experience_required', 'work_experience_years',
            'required_documents', 'additional_notes'
        ]
        
        for col, header in enumerate(country_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Sample country requirements
        sample_countries = [
            ['MSc Data Science', 'University of Manchester', 'PK',
             '16 years education', 'BSc, BE, BS', 3.0, 60, 2.8,
             6.5, 6.0, 6.0, 6.0, 6.0, 90, 62, 115,
             'No', 0, 'Transcripts, Degree Certificate, Passport', 'HEC attested required'],
            ['MSc Data Science', 'University of Manchester', 'BD',
             '4-year Bachelor degree', 'BSc, BEng', 3.0, 55, 2.5,
             6.5, 6.0, 6.0, 6.0, 6.0, 90, 62, 115,
             'No', 0, 'Transcripts, Degree, Passport', ''],
            ['MSc Data Science', 'University of Manchester', 'IN',
             '4-year Bachelor degree', 'BTech, BE, BSc', 3.0, 60, 7.0,
             6.5, 6.0, 6.0, 6.0, 6.0, 90, 62, 115,
             'No', 0, 'Transcripts, Degree, Passport', 'First class preferred'],
            ['MSc Data Science', 'University of Manchester', 'NG',
             "Bachelor's degree", 'BSc', 3.0, 55, 2.5,
             6.5, 6.0, 6.0, 6.0, 6.0, 90, 62, 115,
             'No', 0, 'Transcripts, Degree, Passport', ''],
        ]
        
        for row_num, row_data in enumerate(sample_countries, 2):
            for col, value in enumerate(row_data, 1):
                ws2.cell(row=row_num, column=col, value=value)
        
        # Reference sheet with valid values
        ws3 = wb.create_sheet("Reference")
        ws3['A1'] = 'Level Options'
        levels = ['FOUNDATION', 'UG', 'PG', 'PHD', 'PRE_MASTERS', 'DIPLOMA']
        for i, level in enumerate(levels, 2):
            ws3[f'A{i}'] = level
        
        ws3['B1'] = 'Country Codes'
        countries = ['PK', 'BD', 'IN', 'NG', 'GH', 'KE', 'NP', 'LK', 'AE', 'SA', 'OTHER']
        for i, country in enumerate(countries, 2):
            ws3[f'B{i}'] = country
        
        ws3['C1'] = 'Currency'
        currencies = ['GBP', 'USD', 'EUR']
        for i, curr in enumerate(currencies, 2):
            ws3[f'C{i}'] = curr
        
        # Save
        output_path = 'data/course_import_template.xlsx'
        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(f'Template saved to: {output_path}'))

    def import_courses(self, file_path, university_filter=None, dry_run=False):
        """Import courses from Excel file."""
        try:
            df_courses = pd.read_excel(file_path, sheet_name='Courses')
        except Exception as e:
            self.stderr.write(f'Error reading file: {e}')
            return
        
        # Try to read country requirements sheet
        try:
            df_requirements = pd.read_excel(file_path, sheet_name='Country_Requirements')
            has_requirements = True
        except:
            df_requirements = None
            has_requirements = False
        
        self.stdout.write(f'\n=== Importing {len(df_courses)} courses ===\n')
        
        created = 0
        updated = 0
        errors = []
        
        for idx, row in df_courses.iterrows():
            try:
                university_name = str(row['university_name']).strip()
                
                if university_filter and university_filter.lower() not in university_name.lower():
                    continue
                
                # Find university
                university = University.objects.filter(name__icontains=university_name).first()
                if not university:
                    # Auto-create university if not found (as non-partner)
                    university = University.objects.create(
                        name=university_name,
                        is_partner=False,
                        country='UK'  # Default to UK, can be updated later
                    )
                    self.stdout.write(self.style.WARNING(f"  Auto-created university: {university_name}"))
                
                course_name = str(row['course_name']).strip()
                
                # Parse data
                course_data = {
                    'name': course_name,
                    'level': str(row.get('level', 'PG')).upper(),
                    'duration': str(row.get('duration', '1 Year')),
                    'duration_months': self.safe_int(row.get('duration_months')),
                    'tuition_fee': self.safe_decimal(row.get('tuition_fee_intl', 0)),
                    'tuition_fee_home': self.safe_decimal(row.get('tuition_fee_home')),
                    'currency': str(row.get('currency', 'GBP')).upper(),
                    'deposit_required': self.safe_decimal(row.get('deposit_required')),
                    'intake_january': self.parse_bool(row.get('intake_january')),
                    'intake_may': self.parse_bool(row.get('intake_may')),
                    'intake_september': self.parse_bool(row.get('intake_september', True)),
                    'intakes': str(row.get('intakes', '')) if pd.notna(row.get('intakes')) else '',
                    'ielts_overall': self.safe_decimal(row.get('ielts_overall')),
                    'ielts_each_band': self.safe_decimal(row.get('ielts_each_band')),
                    'academic_requirement': str(row.get('academic_requirement', '')) if pd.notna(row.get('academic_requirement')) else '',
                    'work_experience_required': self.parse_bool(row.get('work_experience_required')),
                    'work_experience_years': self.safe_int(row.get('work_experience_years', 0)),
                    'official_url': str(row.get('official_url', '')) if pd.notna(row.get('official_url')) else '',
                    'department': str(row.get('department', '')) if pd.notna(row.get('department')) else '',
                    'course_code': str(row.get('course_code', '')) if pd.notna(row.get('course_code')) else '',
                    'data_source': 'EXCEL_IMPORT',
                    'last_verified': datetime.now(),
                    'is_data_verified': True,
                }
                
                if dry_run:
                    self.stdout.write(f"  Would create/update: {course_name} at {university.name}")
                    created += 1
                    continue
                
                # Create or update course
                course, is_created = Course.objects.update_or_create(
                    university=university,
                    name=course_name,
                    defaults=course_data
                )
                
                if is_created:
                    created += 1
                    self.stdout.write(f"  Created: {course_name}")
                else:
                    updated += 1
                    self.stdout.write(f"  Updated: {course_name}")
                    
            except Exception as e:
                errors.append(f"Row {idx+2}: {e}")
        
        # Import country requirements
        if has_requirements and df_requirements is not None and not dry_run:
            self.stdout.write(f'\n=== Importing Country Requirements ===\n')
            req_created = 0
            
            for idx, row in df_requirements.iterrows():
                try:
                    course_name = str(row['course_name']).strip()
                    university_name = str(row['university_name']).strip()
                    country_code = str(row['country_code']).strip().upper()
                    
                    # Find course
                    course = Course.objects.filter(
                        name__iexact=course_name,
                        university__name__icontains=university_name
                    ).first()
                    
                    if not course:
                        continue
                    
                    req_data = {
                        'min_qualification': str(row.get('min_qualification', '')),
                        'accepted_qualifications': str(row.get('accepted_qualifications', '')).split(', ') if pd.notna(row.get('accepted_qualifications')) else [],
                        'min_gpa': self.safe_decimal(row.get('min_gpa')),
                        'min_percentage': self.safe_int(row.get('min_percentage')),
                        'min_cgpa': self.safe_decimal(row.get('min_cgpa')),
                        'ielts_overall': self.safe_decimal(row.get('ielts_overall')),
                        'ielts_speaking': self.safe_decimal(row.get('ielts_speaking')),
                        'ielts_writing': self.safe_decimal(row.get('ielts_writing')),
                        'ielts_reading': self.safe_decimal(row.get('ielts_reading')),
                        'ielts_listening': self.safe_decimal(row.get('ielts_listening')),
                        'toefl_score': self.safe_int(row.get('toefl_score')),
                        'pte_score': self.safe_int(row.get('pte_score')),
                        'duolingo_score': self.safe_int(row.get('duolingo_score')),
                        'work_experience_required': self.parse_bool(row.get('work_experience_required')),
                        'work_experience_years': self.safe_int(row.get('work_experience_years', 0)),
                        'required_documents': str(row.get('required_documents', '')).split(', ') if pd.notna(row.get('required_documents')) else [],
                        'additional_notes': str(row.get('additional_notes', '')) if pd.notna(row.get('additional_notes')) else '',
                    }
                    
                    CourseRequirement.objects.update_or_create(
                        course=course,
                        country=country_code,
                        defaults=req_data
                    )
                    req_created += 1
                    
                except Exception as e:
                    errors.append(f"Requirements Row {idx+2}: {e}")
            
            self.stdout.write(f'  Country requirements processed: {req_created}')
        
        # Summary
        self.stdout.write(f'\n=== IMPORT SUMMARY ===')
        self.stdout.write(f'Created: {created}')
        self.stdout.write(f'Updated: {updated}')
        self.stdout.write(f'Errors: {len(errors)}')
        
        if errors:
            self.stdout.write('\nErrors:')
            for err in errors[:10]:
                self.stderr.write(f'  {err}')
            if len(errors) > 10:
                self.stderr.write(f'  ... and {len(errors)-10} more')

    def safe_decimal(self, value):
        """Safely convert to Decimal."""
        if pd.isna(value) or value == '' or value is None:
            return None
        try:
            return Decimal(str(value))
        except:
            return None

    def safe_int(self, value):
        """Safely convert to int."""
        if pd.isna(value) or value == '' or value is None:
            return None
        try:
            return int(float(value))
        except:
            return None

    def parse_bool(self, value):
        """Parse boolean from various formats."""
        if pd.isna(value):
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['yes', 'true', '1', 'y']
        return bool(value)
